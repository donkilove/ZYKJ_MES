import base64
import csv
import io
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_current_active_user, require_permission
from app.db.session import get_db
from app.models.user import User
from app.models.user_session import UserSession
from app.schemas.common import ApiResponse, success_response
from app.schemas.user import (
    UserCreate,
    UserDeleteRequest,
    UserDeleteResult,
    UserExportResult,
    UserExportTaskCreateRequest,
    UserExportTaskItem,
    UserExportTaskListResult,
    UserImportItemResult,
    UserImportResult,
    UserItem,
    UserLifecycleRequest,
    UserLifecycleResult,
    UserListResult,
    UserOnlineStatusResult,
    UserPasswordResetResult,
    UserResetPasswordRequest,
    UserRestoreRequest,
    UserUpdate,
)
from app.services.audit_service import write_audit_log
from app.services.message_service import create_message_for_users
from app.services.session_service import list_online_user_ids
from app.services.user_service import (
    create_user,
    delete_user,
    get_user_by_id,
    list_users,
    reset_user_password,
    restore_user,
    set_user_active,
    update_user,
)
from app.services.user_export_task_service import (
    USER_EXPORT_STATUS_SUCCEEDED,
    create_user_export_task,
    get_user_export_task,
    list_user_export_tasks,
    run_user_export_task,
)


router = APIRouter()


def to_user_export_task_item(task) -> UserExportTaskItem:
    return UserExportTaskItem(
        id=task.id,
        task_code=task.task_code,
        status=task.status,
        format=task.format,
        deleted_scope=task.deleted_scope,
        keyword=task.keyword,
        role_code=task.role_code,
        is_active=task.is_active,
        record_count=task.record_count,
        file_name=task.file_name,
        mime_type=task.mime_type,
        failure_reason=task.failure_reason,
        requested_at=task.requested_at,
        started_at=task.started_at,
        finished_at=task.finished_at,
        expires_at=task.expires_at,
    )


def _resolve_online_user_ids_for_users(
    db: Session,
    users: list[User],
    *,
    preloaded_online_user_ids: set[int] | None = None,
) -> set[int]:
    if not users:
        return set()
    current_page_user_ids = {user.id for user in users}
    if preloaded_online_user_ids is not None:
        return {
            user_id
            for user_id in preloaded_online_user_ids
            if user_id in current_page_user_ids
        }
    return list_online_user_ids(db, candidate_user_ids=list(current_page_user_ids))


def to_user_item(user: User, *, online_user_ids: set[int] | None = None) -> UserItem:
    is_online = user.id in (online_user_ids or set())
    stage_name = user.stage.name if user.stage else None
    primary_role = sorted(user.roles, key=lambda role: role.code)[0] if user.roles else None
    return UserItem(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        remark=user.remark,
        is_online=is_online,
        is_active=user.is_active,
        is_deleted=user.is_deleted,
        must_change_password=user.must_change_password,
        last_seen_at=user.last_login_at,
        stage_id=user.stage_id,
        stage_name=stage_name,
        role_code=primary_role.code if primary_role else None,
        role_name=primary_role.name if primary_role else None,
        last_login_at=user.last_login_at,
        last_login_ip=user.last_login_ip,
        password_changed_at=user.password_changed_at,
        created_at=user.created_at,
        updated_at=user.updated_at,
    )


def _build_csv_export(users: list[User], online_user_ids: set[int]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "id",
            "用户名",
            "角色",
            "工段",
            "在线状态",
            "账号状态",
            "首次登录需改密",
            "创建时间",
            "最近登录时间",
        ]
    )
    for user in users:
        writer.writerow(
            [
                user.id,
                user.username,
                next(iter(sorted(role.name for role in user.roles)), "/"),
                user.stage.name if user.stage else "/",
                "在线" if user.id in online_user_ids else "离线",
                "启用" if user.is_active else "停用",
                "是" if user.must_change_password else "否",
                user.created_at.isoformat(),
                user.last_login_at.isoformat() if user.last_login_at else "",
            ]
        )
    return base64.b64encode(buffer.getvalue().encode("utf-8-sig")).decode("ascii")


def _build_excel_export(users: list[User], online_user_ids: set[int]) -> str:
    try:
        import openpyxl
    except ImportError:
        return ""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "用户列表"
    headers = ["id", "用户名", "角色", "工段", "在线状态", "账号状态", "首次登录需改密", "创建时间", "最近登录时间"]
    ws.append(headers)
    for user in users:
        ws.append([
            user.id,
            user.username,
            next(iter(sorted(role.name for role in user.roles)), "/"),
            user.stage.name if user.stage else "/",
            "在线" if user.id in online_user_ids else "离线",
            "启用" if user.is_active else "停用",
            "是" if user.must_change_password else "否",
            user.created_at.isoformat(),
            user.last_login_at.isoformat() if user.last_login_at else "",
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


@router.get("", response_model=ApiResponse[UserListResult])
def get_users(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    keyword: str | None = Query(default=None),
    role_code: str | None = Query(default=None),
    stage_id: int | None = Query(default=None, ge=1),
    is_active: bool | None = Query(default=None),
    is_online: bool | None = Query(default=None),
    deleted_scope: str = Query(default="active", pattern="^(active|deleted|all)$"),
    include_deleted: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("user.users.list")),
) -> ApiResponse[UserListResult]:
    online_user_ids_for_filter = (
        list_online_user_ids(db) if is_online is not None else None
    )
    total, users = list_users(
        db,
        page=page,
        page_size=page_size,
        keyword=keyword,
        role_code=role_code,
        stage_id=stage_id,
        is_online=is_online,
        online_user_ids=online_user_ids_for_filter,
        is_active=is_active,
        deleted_scope=deleted_scope,
        include_deleted=include_deleted,
    )
    online_user_ids = _resolve_online_user_ids_for_users(
        db,
        users,
        preloaded_online_user_ids=online_user_ids_for_filter,
    )
    result = UserListResult(
        total=total,
        items=[to_user_item(user, online_user_ids=online_user_ids) for user in users],
    )
    return success_response(result)


@router.get("/export", response_model=ApiResponse[UserExportResult])
def export_users(
    keyword: str | None = Query(default=None),
    role_code: str | None = Query(default=None),
    stage_id: int | None = Query(default=None, ge=1),
    is_active: bool | None = Query(default=None),
    is_online: bool | None = Query(default=None),
    deleted_scope: str = Query(default="active", pattern="^(active|deleted|all)$"),
    include_deleted: bool = Query(default=False),
    format: str = Query(default="csv", pattern="^(csv|excel)$"),
    db: Session = Depends(get_db),
    permission_user: User = Depends(require_permission("user.users.export")),
) -> ApiResponse[UserExportResult]:
    _ = permission_user
    online_user_ids_for_filter = (
        list_online_user_ids(db) if is_online is not None else None
    )
    total, users = list_users(
        db,
        page=1,
        page_size=5000,
        keyword=keyword,
        role_code=role_code,
        stage_id=stage_id,
        is_online=is_online,
        online_user_ids=online_user_ids_for_filter,
        is_active=is_active,
        deleted_scope=deleted_scope,
        include_deleted=include_deleted,
    )
    online_user_ids = _resolve_online_user_ids_for_users(
        db,
        users,
        preloaded_online_user_ids=online_user_ids_for_filter,
    )
    _ = total
    if format == "excel":
        content_base64 = _build_excel_export(users, online_user_ids)
        if not content_base64:
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Excel export not available (openpyxl not installed)")
        return success_response(
            UserExportResult(
                filename="users_export.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content_base64=content_base64,
            )
        )
    content_base64 = _build_csv_export(users, online_user_ids)
    return success_response(
        UserExportResult(
            filename="users_export.csv",
            content_type="text/csv",
            content_base64=content_base64,
        )
    )


@router.post("/export-tasks", response_model=ApiResponse[UserExportTaskItem])
def create_user_export_task_api(
    payload: UserExportTaskCreateRequest,
    background_tasks: BackgroundTasks,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.export")),
) -> ApiResponse[UserExportTaskItem]:
    task = create_user_export_task(
        db,
        created_by_user_id=int(current_user.id),
        format=payload.format,
        deleted_scope=payload.deleted_scope,
        keyword=payload.keyword,
        role_code=payload.role_code,
        is_active=payload.is_active,
    )
    write_audit_log(
        db,
        action_code="user.export.create",
        action_name="创建用户导出任务",
        target_type="user_export_task",
        target_id=str(task.id),
        target_name=task.task_code,
        operator=current_user,
        after_data={
            "format": task.format,
            "keyword": task.keyword,
            "role_code": task.role_code,
            "is_active": task.is_active,
            "deleted_scope": task.deleted_scope,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
    )
    db.commit()
    background_tasks.add_task(run_user_export_task, int(task.id))
    return success_response(to_user_export_task_item(task), message="accepted")


@router.get("/export-tasks", response_model=ApiResponse[UserExportTaskListResult])
def list_user_export_tasks_api(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.export")),
) -> ApiResponse[UserExportTaskListResult]:
    tasks = list_user_export_tasks(db, created_by_user_id=int(current_user.id))
    return success_response(
        UserExportTaskListResult(
            total=len(tasks),
            items=[to_user_export_task_item(task) for task in tasks],
        )
    )


@router.get("/export-tasks/{task_id}", response_model=ApiResponse[UserExportTaskItem])
def get_user_export_task_api(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.export")),
) -> ApiResponse[UserExportTaskItem]:
    task = get_user_export_task(
        db,
        task_id=task_id,
        created_by_user_id=int(current_user.id),
    )
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出任务不存在")
    return success_response(to_user_export_task_item(task))


@router.get("/export-tasks/{task_id}/download", response_class=StreamingResponse)
def download_user_export_task_api(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.export")),
) -> StreamingResponse:
    task = get_user_export_task(
        db,
        task_id=task_id,
        created_by_user_id=int(current_user.id),
    )
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出任务不存在")
    if task.status != USER_EXPORT_STATUS_SUCCEEDED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导出任务尚未完成")
    if not task.storage_path or not task.file_name or not task.mime_type:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导出文件不存在，请重新导出")
    file_path = Path(task.storage_path)
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导出文件已过期，请重新导出")
    headers = {
        "Content-Disposition": f'attachment; filename="{task.file_name}"',
        "Content-Type": task.mime_type,
    }
    return StreamingResponse(
        iter([file_path.read_bytes()]),
        media_type=task.mime_type,
        headers=headers,
    )


@router.post("", response_model=ApiResponse[UserItem], status_code=status.HTTP_201_CREATED)
def create_user_api(
    payload: UserCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.create")),
) -> ApiResponse[UserItem]:
    user, error_message = create_user(db, payload)
    if error_message:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_message)
    if not user:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to create user")
    try:
        write_audit_log(
            db,
            action_code="user.create",
            action_name="新建用户",
            target_type="user",
            target_id=str(user.id),
            target_name=user.username,
            operator=current_user,
            after_data={
                "username": user.username,
                "role_code": user.roles[0].code if user.roles else None,
                "stage_id": user.stage_id,
            },
            ip_address=request.client.host if request and request.client else None,
            terminal_info=request.headers.get("user-agent") if request else None,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return success_response(
        to_user_item(
            user,
            online_user_ids=list_online_user_ids(
                db,
                candidate_user_ids=[user.id],
            ),
        ),
        message="created",
    )


@router.get("/online-status", response_model=ApiResponse[UserOnlineStatusResult])
def get_users_online_status(
    user_id: list[int] = Query(default=[]),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("user.users.list")),
) -> ApiResponse[UserOnlineStatusResult]:
    requested_user_ids = sorted({value for value in user_id if value > 0})
    if not requested_user_ids:
        return success_response(UserOnlineStatusResult(user_ids=[]))
    online_user_ids = sorted(
        list_online_user_ids(db, candidate_user_ids=requested_user_ids)
    )
    return success_response(UserOnlineStatusResult(user_ids=online_user_ids))


@router.get("/{user_id}", response_model=ApiResponse[UserItem])
def get_user_detail(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("user.users.detail")),
) -> ApiResponse[UserItem]:
    user = get_user_by_id(db, user_id, include_deleted=True)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    return success_response(
        to_user_item(
            user,
            online_user_ids=list_online_user_ids(
                db,
                candidate_user_ids=[user.id],
            ),
        )
    )


@router.put("/{user_id}", response_model=ApiResponse[UserItem])
def update_user_api(
    user_id: int,
    payload: UserUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.update")),
) -> ApiResponse[UserItem]:
    user = get_user_by_id(db, user_id, include_deleted=True)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    before_data = {
        "username": user.username,
        "full_name": user.full_name,
        "is_active": user.is_active,
        "role_code": user.roles[0].code if user.roles else None,
        "stage_id": user.stage_id,
    }
    updated, error_message = update_user(
        db,
        user=user,
        payload=payload,
        operator=current_user,
    )
    if error_message:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_message)
    if not updated:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to update user")

    write_audit_log(
        db,
        action_code="user.update",
        action_name="编辑用户",
        target_type="user",
        target_id=str(updated.id),
        target_name=updated.username,
        operator=current_user,
        before_data=before_data,
        after_data={
            "username": updated.username,
            "full_name": updated.full_name,
            "is_active": updated.is_active,
            "role_code": updated.roles[0].code if updated.roles else None,
            "stage_id": updated.stage_id,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
    )
    db.commit()
    return success_response(
        to_user_item(
            updated,
            online_user_ids=list_online_user_ids(
                db,
                candidate_user_ids=[updated.id],
            ),
        )
    )


@router.post("/{user_id}/enable", response_model=ApiResponse[UserLifecycleResult])
def enable_user_api(
    user_id: int,
    payload: UserLifecycleRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.enable")),
) -> ApiResponse[UserLifecycleResult]:
    user = get_user_by_id(db, user_id, include_deleted=True)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    before_online_user_ids = list_online_user_ids(db, candidate_user_ids=[user.id])
    before_data = {
        "is_active": user.is_active,
        "is_online": user.id in before_online_user_ids,
        "active_session_count": int(
            db.query(UserSession)
            .filter(
                UserSession.user_id == user.id,
                UserSession.status == "active",
            )
            .count()
        ),
    }
    lifecycle_result, error_message = set_user_active(db, user=user, active=True)
    if error_message:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_message)
    if not lifecycle_result:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to enable user")
    updated, lifecycle_change = lifecycle_result
    write_audit_log(
        db,
        action_code="user.enable",
        action_name="启用用户",
        target_type="user",
        target_id=str(updated.id),
        target_name=updated.username,
        operator=current_user,
        before_data=before_data,
        after_data={
            "is_active": updated.is_active,
            "forced_offline_session_count": lifecycle_change.forced_offline_session_count,
            "cleared_online_status": lifecycle_change.cleared_online_status,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
        remark=(payload.remark or "").strip() or None,
    )
    db.commit()
    return success_response(
        UserLifecycleResult(
            user=to_user_item(
                updated,
                online_user_ids=list_online_user_ids(
                    db,
                    candidate_user_ids=[updated.id],
                ),
            ),
            forced_offline_session_count=lifecycle_change.forced_offline_session_count,
            cleared_online_status=lifecycle_change.cleared_online_status,
        )
    )


@router.post("/{user_id}/disable", response_model=ApiResponse[UserLifecycleResult])
def disable_user_api(
    user_id: int,
    payload: UserLifecycleRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.disable")),
) -> ApiResponse[UserLifecycleResult]:
    user = get_user_by_id(db, user_id, include_deleted=True)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    remark = (payload.remark or "").strip()
    if not remark:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="停用原因不能为空")
    before_online_user_ids = list_online_user_ids(db, candidate_user_ids=[user.id])
    before_data = {
        "is_active": user.is_active,
        "is_online": user.id in before_online_user_ids,
        "active_session_count": int(
            db.query(UserSession)
            .filter(
                UserSession.user_id == user.id,
                UserSession.status == "active",
            )
            .count()
        ),
    }
    lifecycle_result, error_message = set_user_active(db, user=user, active=False)
    if error_message:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_message)
    if not lifecycle_result:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to disable user")
    updated, lifecycle_change = lifecycle_result
    write_audit_log(
        db,
        action_code="user.disable",
        action_name="停用用户",
        target_type="user",
        target_id=str(updated.id),
        target_name=updated.username,
        operator=current_user,
        before_data=before_data,
        after_data={
            "is_active": updated.is_active,
            "forced_offline_session_count": lifecycle_change.forced_offline_session_count,
            "cleared_online_status": lifecycle_change.cleared_online_status,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
        remark=remark,
    )
    db.commit()
    create_message_for_users(
        db,
        message_type="notice",
        priority="important",
        title="账号已被停用",
        summary=f"账号 {updated.username} 已被管理员停用，原因：{remark}。如需恢复请联系系统管理员。",
        content=(
            f"您的账号 {updated.username} 已被管理员 {current_user.username} 停用。"
            f"停用原因：{remark}。如该操作与预期不符，请联系系统管理员核实。"
        ),
        source_module="user",
        source_type="user_disable",
        source_id=str(updated.id),
        source_code=updated.username,
        target_page_code="user",
        target_tab_code="account_settings",
        recipient_user_ids=[updated.id],
        dedupe_key=f"user_disabled_{updated.id}_{int(updated.updated_at.timestamp()) if updated.updated_at else 'now'}",
        created_by_user_id=current_user.id,
    )
    return success_response(
        UserLifecycleResult(
            user=to_user_item(
                updated,
                online_user_ids=list_online_user_ids(
                    db,
                    candidate_user_ids=[updated.id],
                ),
            ),
            forced_offline_session_count=lifecycle_change.forced_offline_session_count,
            cleared_online_status=lifecycle_change.cleared_online_status,
        )
    )


@router.post("/{user_id}/reset-password", response_model=ApiResponse[UserPasswordResetResult])
def reset_password_api(
    user_id: int,
    request: Request,
    payload: UserResetPasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.reset_password")),
) -> ApiResponse[UserPasswordResetResult]:
    user = get_user_by_id(db, user_id, include_deleted=True)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    remark = payload.remark.strip()
    if not remark:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="重置原因不能为空")
    before_online_user_ids = list_online_user_ids(db, candidate_user_ids=[user.id])
    before_data = {
        "is_online": user.id in before_online_user_ids,
        "active_session_count": int(
            db.query(UserSession)
            .filter(
                UserSession.user_id == user.id,
                UserSession.status == "active",
            )
            .count()
        ),
        "must_change_password": user.must_change_password,
        "password_changed_at": user.password_changed_at.isoformat()
        if user.password_changed_at
        else None,
    }
    reset_result, error_message = reset_user_password(
        db,
        user=user,
        new_password=payload.password,
    )
    if error_message:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_message)
    if not reset_result:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to reset password")
    updated, reset_change = reset_result
    write_audit_log(
        db,
        action_code="user.reset_password",
        action_name="重置密码",
        target_type="user",
        target_id=str(updated.id),
        target_name=updated.username,
        operator=current_user,
        before_data=before_data,
        after_data={
            "forced_offline_session_count": reset_change.forced_offline_session_count,
            "cleared_online_status": reset_change.cleared_online_status,
            "must_change_password": reset_change.must_change_password,
            "password_changed_at": updated.password_changed_at.isoformat()
            if updated.password_changed_at
            else None,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
        remark=remark,
    )
    create_message_for_users(
        db,
        message_type="warning",
        priority="important",
        title="密码已被管理员重置",
        summary=(
            f"账号 {updated.username} 的密码已被管理员重置，旧会话已失效，"
            "下次登录必须修改密码。"
        ),
        content=(
            f"您的账号 {updated.username} 已被管理员 {current_user.username} 重置密码。"
            f"重置原因：{remark}。旧会话已失效，下次登录必须先修改密码。"
        ),
        source_module="user",
        source_type="user_reset_password",
        source_id=str(updated.id),
        source_code=updated.username,
        target_page_code="user",
        target_tab_code="account_settings",
        recipient_user_ids=[updated.id],
        dedupe_key=(
            f"user_reset_password_{updated.id}_"
            f"{int(updated.password_changed_at.timestamp()) if updated.password_changed_at else 'now'}"
        ),
        created_by_user_id=current_user.id,
    )
    return success_response(
        UserPasswordResetResult(
            user=to_user_item(
                updated,
                online_user_ids=list_online_user_ids(
                    db,
                    candidate_user_ids=[updated.id],
                ),
            ),
            forced_offline_session_count=reset_change.forced_offline_session_count,
            must_change_password=reset_change.must_change_password,
            cleared_online_status=reset_change.cleared_online_status,
        )
    )


@router.delete("/{user_id}", response_model=ApiResponse[UserDeleteResult])
def delete_user_api(
    user_id: int,
    payload: UserDeleteRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.delete")),
) -> ApiResponse[UserDeleteResult]:
    if user_id == current_user.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot delete current login user")

    user = get_user_by_id(db, user_id, include_deleted=True)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    remark = payload.remark.strip()
    before_online_user_ids = list_online_user_ids(db, candidate_user_ids=[user.id])
    before_data = {
        "username": user.username,
        "role_code": user.roles[0].code if user.roles else None,
        "stage_id": user.stage_id,
        "is_active": user.is_active,
        "is_deleted": user.is_deleted,
        "is_online": user.id in before_online_user_ids,
        "active_session_count": int(
            db.query(UserSession)
            .filter(
                UserSession.user_id == user.id,
                UserSession.status == "active",
            )
            .count()
        ),
    }

    deleted_result, error_message = delete_user(db, user=user)
    if error_message:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_message)
    if not deleted_result:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to delete user")
    deleted, lifecycle_change = deleted_result

    write_audit_log(
        db,
        action_code="user.delete",
        action_name="逻辑删除用户",
        target_type="user",
        target_id=str(deleted.id),
        target_name=deleted.username,
        operator=current_user,
        before_data=before_data,
        after_data={
            "is_deleted": deleted.is_deleted,
            "is_active": deleted.is_active,
            "deleted_at": deleted.deleted_at.isoformat() if deleted.deleted_at else None,
            "forced_offline_session_count": lifecycle_change.forced_offline_session_count,
            "cleared_online_status": lifecycle_change.cleared_online_status,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
        remark=remark,
    )
    db.commit()
    return success_response(
        UserDeleteResult(
            user=to_user_item(
                deleted,
                online_user_ids=list_online_user_ids(
                    db,
                    candidate_user_ids=[deleted.id],
                ),
            ),
            forced_offline_session_count=lifecycle_change.forced_offline_session_count,
            cleared_online_status=lifecycle_change.cleared_online_status,
            deleted=True,
        ),
        message="deleted",
    )


@router.post("/{user_id}/restore", response_model=ApiResponse[UserLifecycleResult])
def restore_user_api(
    user_id: int,
    payload: UserRestoreRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.restore")),
) -> ApiResponse[UserLifecycleResult]:
    user = get_user_by_id(db, user_id, include_deleted=True)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    remark = payload.remark.strip()
    before_data = {
        "username": user.username,
        "role_code": user.roles[0].code if user.roles else None,
        "stage_id": user.stage_id,
        "is_active": user.is_active,
        "is_deleted": user.is_deleted,
        "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
    }

    restored_result, error_message = restore_user(db, user=user)
    if error_message:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_message)
    if not restored_result:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to restore user")
    restored, lifecycle_change = restored_result

    write_audit_log(
        db,
        action_code="user.restore",
        action_name="恢复用户",
        target_type="user",
        target_id=str(restored.id),
        target_name=restored.username,
        operator=current_user,
        before_data=before_data,
        after_data={
            "is_deleted": restored.is_deleted,
            "is_active": restored.is_active,
            "deleted_at": restored.deleted_at.isoformat() if restored.deleted_at else None,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
        remark=remark,
    )
    db.commit()
    return success_response(
        UserLifecycleResult(
            user=to_user_item(
                restored,
                online_user_ids=list_online_user_ids(
                    db,
                    candidate_user_ids=[restored.id],
                ),
            ),
            forced_offline_session_count=lifecycle_change.forced_offline_session_count,
            cleared_online_status=lifecycle_change.cleared_online_status,
        )
    )


@router.post("/import", response_model=ApiResponse[UserImportResult])
async def import_users(
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("user.users.import")),
) -> ApiResponse[UserImportResult]:
    content = await file.read()
    filename = (file.filename or "").lower()
    rows_data: list[dict[str, str]] = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows_data.append({k.strip(): (v or "").strip() for k, v in row.items()})
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Excel文件无有效工作表",
                )
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h or "").strip() for h in next(rows_iter)]
            for row in rows_iter:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = str(val or "").strip()
                rows_data.append(row_dict)
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="服务端未安装openpyxl，无法解析Excel文件",
            )
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="仅支持 .csv、.xlsx、.xls 格式",
        )

    required_columns = {"username", "role_code"}
    if not rows_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="文件为空",
        )
    actual_columns = set(rows_data[0].keys())
    missing = required_columns - actual_columns
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"缺少必须列：{', '.join(sorted(missing))}",
        )

    results: list[UserImportItemResult] = []
    success_count = 0
    failure_count = 0

    for idx, row_data in enumerate(rows_data, start=2):
        username = row_data.get("username", "").strip()
        role_code = row_data.get("role_code", "").strip()
        full_name = row_data.get("full_name", "").strip() or None
        stage_id_raw = row_data.get("stage_id", "").strip()
        stage_id = int(stage_id_raw) if stage_id_raw.isdigit() else None
        remark = row_data.get("remark", "").strip() or None

        if not username:
            results.append(UserImportItemResult(
                row_number=idx, username=username, success=False, error="用户名为空",
            ))
            failure_count += 1
            continue

        payload = UserCreate(
            username=username,
            full_name=full_name,
            password="123456",
            role_code=role_code,
            stage_id=stage_id,
            is_active=True,
        )
        user, error = create_user(db, payload)
        if error:
            results.append(UserImportItemResult(
                row_number=idx, username=username, success=False, error=error,
            ))
            failure_count += 1
            continue

        db.flush()
        results.append(UserImportItemResult(
            row_number=idx, username=username, success=True,
            user_id=user.id if user else None,
        ))
        success_count += 1

    write_audit_log(
        db,
        action_code="user.import",
        action_name="批量导入用户",
        target_type="user",
        target_id=str(current_user.id),
        target_name=current_user.username,
        operator=current_user,
        after_data={
            "total_rows": len(rows_data),
            "success_count": success_count,
            "failure_count": failure_count,
        },
        ip_address=request.client.host if request and request.client else None,
        terminal_info=request.headers.get("user-agent") if request else None,
    )
    db.commit()

    return success_response(
        UserImportResult(
            total_rows=len(rows_data),
            success_count=success_count,
            failure_count=failure_count,
            items=results,
        ),
        message="completed",
    )


@router.get("/import-template")
def download_import_template(
    format: str = Query(default="csv", pattern="^(csv|excel)$"),
    _: User = Depends(require_permission("user.users.import")),
) -> ApiResponse[UserExportResult]:
    headers = ["username", "full_name", "role_code", "stage_id", "remark"]
    sample_rows = [
        ["zhangsan", "张三", "operator", "", "冲压工段"],
        ["lisi", "李四", "production_admin", "", ""],
    ]

    if format == "excel":
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("用户导入模板")
            else:
                ws.title = "用户导入模板"
            ws.append(headers)
            for row in sample_rows:
                ws.append(row)
            buffer = io.BytesIO()
            wb.save(buffer)
            content_base64 = base64.b64encode(buffer.getvalue()).decode()
            return success_response(UserExportResult(
                filename="user_import_template.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content_base64=content_base64,
            ))
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="服务端未安装openpyxl，无法生成Excel模板",
            )

    buffer = io.StringIO()
    buffer.write("\ufeff")
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(sample_rows)
    content_base64 = base64.b64encode(buffer.getvalue().encode("utf-8")).decode()
    return success_response(UserExportResult(
        filename="user_import_template.csv",
        content_type="text/csv",
        content_base64=content_base64,
    ))
